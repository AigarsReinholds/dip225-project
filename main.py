from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import StaleElementReferenceException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service = service, options = option)

url = "https://www.cv.lv/"
driver.get(url)

# pamata filtri
locations = ['Rīga','Iespēja strādāt attālināti']
category = "Informācijas tehnoloģijas"
keywords = ['frontend']
# papildus filtri
minimal_salary = 950
typeof_job = 'Pilna slodze'
sort_results = "Parādīt jaunākos vispirms"

# atrasanas vieta
location_input = driver.find_element(By.CLASS_NAME, "css-bg1rzq-control")
location_input.click()
for location in locations:
  location_input_2 = driver.find_element(By.ID, "react-select-2-input")
  location_input_2.send_keys(location)
  location_input_3 = driver.find_element(By.XPATH, "//div[contains(@class, 'css-dpec0i-option')]")
  location_input_3.click()
# darba kategorija
category_input = driver.find_elements(By.CLASS_NAME, "css-bg1rzq-control")[0]
category_input.click()
category_input_2 = driver.find_element(By.ID, "react-select-3-input")
category_input_2.send_keys(category)
category_input_3 = driver.find_element(By.XPATH, "//div[contains(@class, 'css-dpec0i-option')]")
category_input_3.click()
# atslegvardi
keyword_input = driver.find_elements(By.CLASS_NAME, "css-bg1rzq-control")[1]
keyword_input.click()
for keyword in keywords:
  capitalize_keyword = keyword.capitalize()
  keyword_input_2 = driver.find_elements(By.ID, "react-select-2-input")[1]
  keyword_input_2.send_keys(capitalize_keyword)
  keyword_input_3 = driver.find_element(By.CLASS_NAME, "css-dpec0i-option")
  keyword_input_3.click()
# poga papildus filtriem
button_filters = driver.find_element(By.XPATH, "//span[contains(@class,'jsx-2818744897') and contains(@class, 'search-form__additional-toggle')]")
button_filters.click()
time.sleep(1)
# minimala alga
minimal_salary_input = driver.find_element(By.XPATH, "//input[contains(@name, 'salaryFrom')]")
minimal_salary_input.click()
minimal_salary_input.send_keys(minimal_salary)
# darba veids
typeof_job_input = driver.find_elements(By.CLASS_NAME, "css-1pcexqc-container")[4]
typeof_job_input.click()
typeof_job_input_2 = driver.find_element(By.ID, "react-select-5-input")
typeof_job_input_2.send_keys(typeof_job)
typeof_job_input_3 = driver.find_element(By.XPATH, "//div[contains(@class, 'css-dpec0i-option')]")
typeof_job_input_3.click()
# poga rezultatu radisanai
span_results = driver.find_element(By.XPATH, "//span[contains(@class, 'jsx-2818744897') and contains(@class, 'search-form-footer__item')]")
button_results = span_results.find_element(By.XPATH, "./descendant::button")
if button_results.is_enabled():
  button_results.click()
else:
  print("0 rezultāti")
time.sleep(1)
# rezultatu (vaicajumu) daudzuma 'limit' mainisana
opened_url = driver.current_url
opened_url = opened_url.replace("limit=20","limit=200")
driver.get(opened_url)
time.sleep(1)
# rezultatu kartosana
if(sort_results != ""):
  sort_options = driver.find_element(By.XPATH, "//span[contains(@class, 'jsx-1778535529') and contains(@class, 'select')]")
  sort_options.click()
  time.sleep(1)
  button_sort_newest = driver.find_elements(By.XPATH, "//li[contains(@class, 'jsx-1778535529') and contains(@class,'select__item')]")[1]
  button_sort_newest_2 = button_sort_newest.find_element(By.XPATH, ".//descendant::span")
  button_sort_newest_2.click()
# rezultatu saglabasana
wb = load_workbook("result.xlsx")
ws = wb.active
max_row = ws.max_row
# izdzes info, kas saglabata iepriekseja palaisanas reize
ws.delete_rows(idx = 2, amount = max_row - 1)
time.sleep(1)
max_row = ws.max_row 
job_list_items = driver.find_elements(By.CLASS_NAME, "vacancies-list__item")
for job_item in job_list_items:
  try:
    # izgust info no visiem sludinajumiem
    title = job_item.find_element(By.XPATH, ".//span[contains(@class, 'vacancy-item__title')]").get_attribute("textContent")
    employer_name = job_item.find_element(By.XPATH, ".//div[contains(@class, 'jsx-3024910437') and contains(@class, 'vacancy-item__column')]").get_attribute("textContent")
    job_location = job_item.find_element(By.XPATH, ".//div[contains(@class, 'jsx-3024910437') and contains(@class, 'vacancy-item__locations')]").get_attribute("textContent")
    salary_range = job_item.find_element(By.XPATH, ".//span[contains(@class, 'jsx-3024910437') and contains(@class, 'vacancy-item__salary-label')]").get_attribute("textContent")
    dates_info = job_item.find_element(By.XPATH, ".//div[contains(@class, 'jsx-3024910437') and contains(@class, 'vacancy-item__info-secondary')]")
    dates_info_2 = dates_info.find_element(By.XPATH, ".//span[not(contains(@class, 'vacancy-item__expiry'))]").get_attribute("textContent")
    dates_info_2_space = dates_info_2.replace('Beidzas', ' | Beidzas')
    # iziet cauri katrai sludinajuma lapai
    job_links = []
    job_list_items_links = job_item.find_elements(By.XPATH, ".//a[contains(@class, 'jsx-3024910437') and contains(@class, 'vacancy-item')]")
    for job_item_link in job_list_items_links:
      job_link = job_item_link.get_attribute("href")
      job_links.append(job_link)
    for job_item_link in job_links:
      driver.get(job_item_link)
      ws['H'+str(max_row + 1)].hyperlink = job_item_link 
      job_description_text = ""
      try:
        full_job_description = WebDriverWait(driver, 2.5).until(
          EC.presence_of_all_elements_located(
            (By.XPATH, ".//div[contains(@class, 'jsx-2212276015') and contains(@class, 'vacancy-details__section')]")
          )
        )
        for job_description in full_job_description:
          job_description_text += job_description.text + "\n"
        
      except:
        # nevar izgut info, piemeram, no bildes vai personalizetas lapas
        pass

      driver.back()
      time.sleep(0.5)

    for i in range(len(job_list_items)):
      i += 1
      ws['A'+str(i + 1)] = i
    ws['B'+str(max_row + 1)] = title
    ws['C'+str(max_row + 1)] = employer_name
    ws['D'+str(max_row + 1)] = job_location
    ws['E'+str(max_row + 1)] = salary_range
    ws['G'+str(max_row + 1)] = dates_info_2_space
    ws['F'+str(max_row + 1)] = job_description_text
    ws['F'+str(max_row + 1)].alignment = Alignment(wrap_text=True, vertical='top')
    if job_description_text != "":
      ws.row_dimensions[max_row + 1].height = 300
    else:
      ws.row_dimensions[max_row + 1].height = 30
    ws.column_dimensions['F'].width = 150
    max_row += 1
    
  except StaleElementReferenceException:
    pass

driver.close()
wb.save("result.xlsx")
wb.close()
time.sleep(5)
driver.quit()